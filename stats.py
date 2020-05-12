from openpyxl import Workbook, load_workbook
import statistics

wb = Workbook()
sheet = wb.active
sheets = ["cambridge_scores.xlsx", "harvard_scores.xlsx", "nyu_scores.xlsx", "yale_scores.xlsx", "stanford_scores.xlsx","control_scores.xlsx"]

def calculate(fname):
    wbin = load_workbook(filename=fname)
    print("got here")
    wbinsh = wbin.active

    lists = {}
    lists["Flesh Reading Ease"] = [] 
    lists["Flesh Kincaid Grade Level"] = []
    lists["Coleman Liau Index"] = []
    lists["Gunning Fog Index"] = []
    lists["SMOG Index"] = []
    lists["ARI Index"] = []
    lists["LIX Index"] = []
    lists["Dale-Chall Score"] = []
    lists["TTR Simple"] = []

    j = 2
    
    # while(wbinsh["A"+str(j)].value):
    #     for i in range(9):
    #         lists[wbinsh[chr(65+i) + "1"].value].append(wbinsh[chr(65+i)+str(j)].value)
    #     j+=1
    for j in range(2,200000):
        if wbinsh["A"+str(j)].value:
            for i in range(9):
                lists[wbinsh[chr(65+i) + "1"].value].append(wbinsh[chr(65+i)+str(j)].value)
            # j+=1
    medians = {}
    means = {}
    stdivs = {}
    print(lists)
    for key in lists:
       
        medians[key] = statistics.median(lists[key])
        means[key] = statistics.mean(lists[key])
        stdivs[key] = statistics.stdev(lists[key])

    return({"medians":medians, "means":means, "stdevs":stdivs })


for sheetr in sheets:
    data = calculate(sheetr)
    print(sheetr)
    print()
    for i in data["medians"]:
        print(data["medians"][i] )
    print()
    for i in data["means"]:
        print(data["means"][i] )
    print()
    for i in data["stdevs"]:
        print(data["stdevs"][i] )
    print("===============================")