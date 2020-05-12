from readcalc import readcalc
from openpyxl import Workbook, load_workbook
from lexical_diversity import lex_div as ld #pip install lexical-diversity
from nltk.tokenize import word_tokenize

wb = Workbook()
sheet = wb.active
wbin = load_workbook(filename="yale_tweets.xlsx")
wbinsh = wbin.active
sheet["A1"] = "Flesh Reading Ease"
sheet["B1"] = "Flesh Kincaid Grade Level"
sheet["C1"] = "Coleman Liau Index"
sheet["D1"] = "Gunning Fog Index"
sheet["E1"] = "SMOG Index"
sheet["F1"] = "ARI Index"
sheet["G1"] = "LIX Index"
sheet["H1"] = "Dale-Chall Score"
sheet["I1"] = "TTR Simple"


for i in range(1,144332):
    if len(str(wbinsh["K"+str(i)].value).split(" "))> 15:
        calc = readcalc.ReadCalc(wbinsh["K"+str(i)].value)
        tokenized = word_tokenize(str(wbinsh["K"+str(i)].value))
        sheet["A"+str(i+1)] = calc.get_flesch_reading_ease()
        sheet["B"+str(i+1)] = calc.get_flesch_kincaid_grade_level()
        sheet["C"+str(i+1)] = calc.get_coleman_liau_index()
        sheet["D"+str(i+1)] = calc.get_gunning_fog_index()
        sheet["E"+str(i+1)] = calc.get_smog_index()
        sheet["F"+str(i+1)] = calc.get_ari_index()
        sheet["G"+str(i+1)] = calc.get_lix_index()
        sheet["H"+str(i+1)] = calc.get_dale_chall_score()
        sheet["I"+str(i+1)] = ld.ttr(tokenized)

wb.save(filename="yale_scores.xlsx")